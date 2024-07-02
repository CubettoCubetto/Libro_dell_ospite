import shutil
import login
import pandas as pd
from flask import Flask, jsonify, request, send_file, redirect
from flask_cors import CORS
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook
import pytz

from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)  # This will enable CORS for all routes

estensioni_immagine_supportate = [".jpg", ".png", ".jpeg", ".svg"]


@app.route('/receive/<nomeAccount>', methods=['POST'])
def receive(nomeAccount):
    # Extract the JSON data from the request
    data = request.get_json()
    # Ensure that the JSON contains 'name' and 'message'
    if 'message' not in data:
        return jsonify({'error': 'Invalid input'}), 400

    # check that nomeAccount is valid:
    if nomeAccount == "":
        return jsonify({'error': 'Invalid URL: not specified message destination'}), 400

    if login.login(nomeAccount, "password", justCheckIfUsernameExist=True) != "username esiste":
        raise ValueError(f"Commento inviato a un username non esistente: {nomeAccount}")

    # Get the name and message from the JSON data
    name = data['name']
    message = data['message']
    rating = data["ratings"]
    # Imposta il fuso orario per l'Italia
    italy_tz = pytz.timezone('Europe/Rome')

    # Ottieni l'ora corrente nel fuso orario italiano
    now = datetime.now(italy_tz)
    giorno = now.strftime("%d/%m/%Y")
    orario = now.strftime("%H:%M:%S")
    # Print the name and message to the console
    print(f'Ricevuto un nuovo messaggio da {name} per {nomeAccount}: "{message}" in data {giorno} alle {orario}')

    saveMessage(name, nomeAccount, message, rating, giorno, orario)
    # Respond with a success message
    return jsonify({'status': 'success', 'message': 'Data received'}), 200


@app.route('/get_commenti', methods=['POST'])
def download_commenti():
    data = request.json
    username = data['username']
    password = data['password']

    result = login.login(username, password)
    if result == "successo":
        return send_file(os.path.join("database_commenti", username, f"{username}.xlsx"))
    else:
        return jsonify({"error": result})


@app.route('/get_image/<username>')
def download_image(username):
    for estensione in estensioni_immagine_supportate:
        path = os.path.join("loghi", f"{username}{estensione}")
        if os.path.exists(path):
            return send_file(path)

    # se non trovo niente ne restituisco una alzando un errore in automatico
    return send_file(os.path.join("loghi", f"{username}.png"))


# url shortener
@app.route('/sh/<username>')
def redirect_to_send_comment(username):
    return redirect(
        "https://cubettocubetto.github.io/Libro_dell_ospite/invia_commento/invia_commento.html?username=" + username,
        code=302)


@app.route('/get_params/<username>')
def get_params_question(username):
    # Path to the Excel file
    path = "login.xlsx"

    df = pd.read_excel(path)

    # Filter the dataframe to find the row where the first column matches the username
    user_row = df[df['username'] == username]

    if user_row.empty:
        return jsonify({"error": "User not found"}), 404

    # Select the relevant columns and convert to dictionary
    user_data = user_row.iloc[:, 2:-1].to_dict(orient='records')[0]

    # Flatten the dictionary and convert to a list of key-value pairs to maintain order
    ordered_data = [(key, value) for key, value in user_data.items()]
    print(ordered_data)
    return jsonify(ordered_data)


@app.route('/create_account', methods=['POST'])
def create_account():

    # example of data received
    # {'text1': 'PizzaProva', 'text3': "Com'era la nostra pizza?", 'star1': 'Ingredienti',
    # 'star2': 'Servizio', 'star3': 'Location', 'star4': 'NO', 'star5': 'NO',
    # 'name': 'NO', 'password': '123', 'providedPassword': 'oScjQ0gFyP4jI2R'}
    form_data = {}
    for key in request.form:
        form_data[key] = request.form[key]
        if request.form[key] == "":
            return jsonify({"status": "error", "message": f"{key} was empty, please insert something everywhere"}), 400
    print(form_data)

    # check if the password gived by us is right
    login_result = login.check_password_in_excel(form_data["providedPassword"], form_data["username"])
    if login_result[0] is not True:
        return jsonify({"status": "error", "message": login_result[1]}), 400

    # Salvo il logo
    if 'formFile' in request.files:
        file = request.files['formFile']
        if file:
            format_file = "." + file.filename.split(".")[-1]
            print(format_file)
            if format_file in estensioni_immagine_supportate:
                file_path = os.path.join("loghi", form_data["username"] + format_file)
                file.save(file_path)  # Save the file
            else:
                return jsonify({"status": "error", "message": "Image exstension not found"}), 400
        else:
            return jsonify({"status": "error", "message": "Image not found"}), 400
    login.create_account(form_data)

    print(form_data)  # Print form data to the console

    # Return a JSON response
    return jsonify({"status": "success", "username": form_data["username"]}), 200


def add_data_to_excel(name, message, rating, data, ora, excel_file_name):
    excel_path = os.path.join("database_commenti", excel_file_name, f"{excel_file_name}.xlsx")
    try:
        try:
            # Load the workbook (create if not existing)
            wb = openpyxl.load_workbook(excel_path)
        except FileNotFoundError:
            # If the file does not exist, create a new workbook
            wb = Workbook()

        # Select the active sheet (first sheet by default)
        sheet = wb.active

        # Determine the next available row
        next_row = sheet.max_row + 1

        # Add data to the sheet
        sheet[f'A{next_row}'] = name
        sheet[f'B{next_row}'] = message

        sheet[f'C{next_row}'] = rating[0]
        sheet[f'D{next_row}'] = rating[1]
        sheet[f'E{next_row}'] = rating[2]
        sheet[f'F{next_row}'] = rating[3]
        sheet[f'G{next_row}'] = rating[4]

        sheet[f'H{next_row}'] = data
        sheet[f'I{next_row}'] = ora

        # Save the workbook
        wb.save(excel_path)
        print(f"Data added successfully to {excel_path}: Name='{name}', ID='{message}'")

    except Exception as e:
        print("an error occured while modifying orignal excel file: ", e)


def saveMessage(name, nomeAccount, message, rating, giorno, orario):
    # file di testo
    da_scrivere = f"{name} alle {orario} ha lasciato i seguenti voti: {rating}, e ha scritto:\n{message}\n\n"

    # check if the folder with the name already exists, else create it
    folder_path = os.path.join("database_commenti", nomeAccount)
    if not os.path.exists(folder_path):
        os.mkdir(folder_path)

    try:
        with open(os.path.join("database_commenti", nomeAccount, giorno.replace("/", "-")) + ".txt", 'a') as file:
            file.write(da_scrivere)
    except IOError as e:
        print("Error while writing to the file:", e)

    # file excel
    try:
        add_data_to_excel(name, message, rating, giorno, orario, nomeAccount)
    except Exception as e:
        print(f"errore modificando il file originale excel: {e}")


if __name__ == '__main__':
    app.run(debug=True, port=8013)
