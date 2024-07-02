import pandas as pd
import random
import string
import os
from openpyxl import Workbook
from openpyxl import load_workbook

login_file_path = "login.xlsx"
passwords_file_path = "passwords.xlsx"


def login(username, password, justCheckIfUsernameExist=False) -> str:
    # check that I am not receiving an hacker attack to try to break into the database...
    if username == "username":
        return "Username non valido, mi spiace"

    # Read the Excel file
    try:
        df = pd.read_excel(login_file_path)
    except FileNotFoundError:
        return "Errore interno, file non trovato"
    except Exception as e:
        return "Errore interno, errore generico"

    # Check if the provided username exists
    user_row = df[df['username'] == username]
    if justCheckIfUsernameExist and not user_row.empty:
        return "username esiste"

    if user_row.empty:
        return "Username non esistente"

    # Check if the provided password matches the username's password
    if user_row.iloc[0]['password'] == password:
        return "successo"
    else:
        return "password sbagliata"


def create_account(form_data):
    print("Creating account with data:", form_data)
    EXCEL_FILE = 'login.xlsx'
    # Check if the file exists
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Create headers if the file is new
        headers = list(form_data.keys())
        ws.append(headers)

    # Check if username already exists in the Excel file
    usernames = [cell.value for cell in ws['A']]  # Column A contains usernames
    if form_data['username'] in usernames:
        # If username exists, find the row index and overwrite the data
        row_index = usernames.index(form_data['username']) + 1  # +1 because index is zero-based
        for col_index, key in enumerate(form_data.keys(), start=1):
            ws.cell(row=row_index, column=col_index).value = form_data[key]
    else:
        # Add the form data as a new row
        ws.append(list(form_data.values()))

    # Save the workbook
    wb.save(EXCEL_FILE)


def generate_random_code(length=15):
    """Generate a random alphanumeric code of specified length."""
    characters = string.ascii_letters + string.digits
    return ''.join(random.choice(characters) for _ in range(length))


def create_excel_with_random_codes():
    """Create an Excel file with 50 rows of random codes and boolean values."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Random Codes'

    # Write headers
    sheet['A1'] = 'Password'
    sheet['B1'] = 'Chi la sta usando?'
    sheet['C1'] = 'Expired?'
    # Generate 50 rows of data
    for row in range(2, 52):  # 2 to 51 (inclusive)
        random_code = generate_random_code()
        boolean_value = False

        sheet.cell(row=row, column=1).value = random_code
        sheet.cell(row=row, column=2).value = False
        sheet.cell(row=row, column=3).value = False

    # Save the workbook
    wb.save(passwords_file_path)


def check_password_in_excel(password_to_check, username) -> tuple:
    """Check if a code exists in the Excel file and if its corresponding boolean value is the username given."""
    wb = load_workbook(passwords_file_path)
    sheet = wb.active

    # Iterate through rows to find the code
    for row in range(2, sheet.max_row + 1):  # Start from 2nd row, since 1st row is header
        cell_value = sheet.cell(row=row, column=1).value
        user = sheet.cell(row=row, column=2).value
        expired = sheet.cell(row=row, column=3).value
        if cell_value == password_to_check:
            if user == username or user is False:
                if not expired:
                    # assegno questa password a questo utente
                    sheet.cell(row=row, column=2).value = username
                    wb.save(passwords_file_path)
                    return True, "success"
                else:
                    return False, "Account expired, please contact us to renew the account"
            else:
                return False, "wrong username or password not validated"

    # Code not found in the Excel file
    return False, "password fornita da noi non valida, controlla che sia giusta"


if __name__ == "__main__":
    if input("SICURO DI VOLER SOVVRASCRIVERE TUTTE LE PASSWORD PER CREARE ACCOUNT? SI/NO") == "SI":
        create_excel_with_random_codes()
